import smtplib
import openpyxl
import sys
wb = openpyxl.load_workbook('thelist.xlsx')
journalists = {}
sheet = wb.get_sheet_by_name('Лист1')
for r in range(1, sheet.max_row + 1):
    name = sheet.cell(row=r, column=1).value
    email = sheet.cell(row=r, column=2).value
    journalists[name] = email
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('example@example.com', 'password')
for name, email in journalists.items():
    body = "Subject: hello.\nDear %s,\n<b>This</b> is a test." % name
    print('Отсылаю на адрес %s...' % email)
    sendstatus = smtpObj.sendmail('kasoev.gr@phystech.edu', email, body)
    if sendstatus != {}:
        print('Ошибка при отправке на %s: %s' % email, sendstatus)
    else:
        print('Успешно выслано на %s' % email)
smtpObj.quit()
